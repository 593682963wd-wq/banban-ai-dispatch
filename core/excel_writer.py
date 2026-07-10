"""分配结果导出为 Excel。

输出两张表：
  - 「分配明细」：每架飞机所属席位 + 航段明细
  - 「席位汇总」：两席位对比 + 均衡指标
席位 1 用浅灰、席位 2 用浅蓝填充，与公司系统的灰/蓝勾选保持一致。
"""
from __future__ import annotations

import io

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from .airports import airport_name
from .models import AllocationResult

SEAT1_FILL = PatternFill("solid", fgColor="E7E9EC")   # 浅灰 → 席位1
SEAT2_FILL = PatternFill("solid", fgColor="BDD7EE")   # 浅蓝 → 席位2
HEAD_FILL = PatternFill("solid", fgColor="1F4E79")
HEAD_FONT = Font(color="FFFFFF", bold=True)
CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
THIN = Side(style="thin", color="BFBFBF")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


def _style_header(ws, row, ncols):
    for c in range(1, ncols + 1):
        cell = ws.cell(row=row, column=c)
        cell.fill = HEAD_FILL
        cell.font = HEAD_FONT
        cell.alignment = CENTER
        cell.border = BORDER


def _autosize(ws, widths):
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w


def build_workbook(result: AllocationResult) -> Workbook:
    wb = Workbook()
    by_tail = {ac.tail: ac for ac in result.aircrafts}

    # ── Sheet 1：分配明细 ──
    ws = wb.active
    ws.title = "分配明细"
    headers = ["所属席位", "机号", "机型", "航班数", "空任务", "时段A", "时段B",
               "C类", "B类", "长沙出港", "讲解量", "过夜地", "航段明细"]
    ws.append(headers)
    _style_header(ws, 1, len(headers))

    def _legs_text(ac):
        parts = []
        for f in ac.flights:
            parts.append(f"{f.dep_hhmm} {airport_name(f.dep_icao)}→{airport_name(f.arr_icao)}")
        return " ; ".join(parts)

    for seat, fill in ((result.seat1, SEAT1_FILL), (result.seat2, SEAT2_FILL)):
        for tail in seat.tails:
            ac = by_tail.get(tail)
            if ac is None:
                continue
            a, b = ac.n_segment(
                result.config.split_minutes,
                result.config.segment_start_minutes,
                result.config.segment_end_minutes,
                result.config.service_date,
            )
            ov = ac.overnight_dest
            ws.append([
                seat.name, ac.tail, ac.ac_type, ac.n_flights,
                "是" if ac.n_flights == 0 else "", a, b,
                ac.n_c_class, ac.n_b_class, ac.n_changsha_dep, ac.n_briefing,
                airport_name(ov) if ov else "",
                _legs_text(ac) if ac.n_flights else "空任务/停场/备用",
            ])
            for c in range(1, len(headers) + 1):
                cell = ws.cell(row=ws.max_row, column=c)
                cell.fill = fill
                cell.border = BORDER
                cell.alignment = Alignment(
                    horizontal="left" if c == len(headers) else "center",
                    vertical="center",
                )

    _autosize(ws, [11, 9, 16, 7, 8, 7, 7, 6, 6, 10, 8, 9, 70])
    ws.freeze_panes = "A2"

    # ── Sheet 2：席位汇总 ──
    ws2 = wb.create_sheet("席位汇总")
    ws2.append(["指标", "放行席位1", "放行席位2", "差值"])
    _style_header(ws2, 1, 4)
    rows = [
        ("飞机数", len(result.seat1.tails), len(result.seat2.tails)),
        ("空任务飞机", result.seat1.n_idle_aircraft, result.seat2.n_idle_aircraft),
        ("航班任务数", result.seat1.n_flights, result.seat2.n_flights),
        ("时段A航班", result.seat1.n_seg_a, result.seat2.n_seg_a),
        ("时段B航班", result.seat1.n_seg_b, result.seat2.n_seg_b),
        ("C类机场航班", result.seat1.n_c_class, result.seat2.n_c_class),
        ("B类机场航班", result.seat1.n_b_class, result.seat2.n_b_class),
        ("长沙出港航班", result.seat1.n_changsha_dep, result.seat2.n_changsha_dep),
        ("讲解量", result.seat1.n_briefing, result.seat2.n_briefing),
    ]
    for name, v1, v2 in rows:
        ws2.append([name, v1, v2, abs(v1 - v2)])
    for r in range(2, ws2.max_row + 1):
        ws2.cell(row=r, column=2).fill = SEAT1_FILL
        ws2.cell(row=r, column=3).fill = SEAT2_FILL
        for c in range(1, 5):
            ws2.cell(row=r, column=c).alignment = CENTER
            ws2.cell(row=r, column=c).border = BORDER
    _autosize(ws2, [16, 12, 12, 8])

    # 均衡指标
    ws2.append([])
    start = ws2.max_row + 1
    ws2.cell(row=start, column=1, value="均衡指标（越小越均衡）")
    ws2.cell(row=start, column=1).font = Font(bold=True)
    for name, val in result.metrics.items():
        ws2.append([name, val])
        ws2.cell(row=ws2.max_row, column=1).border = BORDER
        ws2.cell(row=ws2.max_row, column=2).border = BORDER

    # ── Sheet 3：席位机号清单（便于直接抄）──
    ws3 = wb.create_sheet("席位机号清单")
    ws3.append(["放行席位1", "放行席位2"])
    _style_header(ws3, 1, 2)
    t1, t2 = result.seat1.tails, result.seat2.tails
    for i in range(max(len(t1), len(t2))):
        ws3.append([
            t1[i] if i < len(t1) else "",
            t2[i] if i < len(t2) else "",
        ])
        ws3.cell(row=ws3.max_row, column=1).fill = SEAT1_FILL
        ws3.cell(row=ws3.max_row, column=2).fill = SEAT2_FILL
        for c in (1, 2):
            ws3.cell(row=ws3.max_row, column=c).alignment = CENTER
            ws3.cell(row=ws3.max_row, column=c).border = BORDER
    _autosize(ws3, [16, 16])

    return wb


def write_allocation_excel(result: AllocationResult, path: str) -> str:
    wb = build_workbook(result)
    wb.save(path)
    return path


def build_excel_bytes(result: AllocationResult) -> bytes:
    wb = build_workbook(result)
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
